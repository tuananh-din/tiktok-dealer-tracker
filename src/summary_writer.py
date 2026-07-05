"""Build a per-dealer summary from the full video report.

Columns: Đại lý | Số video | Tổng view | Tổng tym | Tổng comment | Engagement %
Engagement % = (likes + comments) / views * 100, aggregated per dealer.
"""
import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook

HEADERS = ["Đại lý", "Số video", "Tổng view", "Tổng tym", "Tổng comment", "Engagement %"]


def _num(v) -> int:
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0


def build_rows(excel_file: Path) -> list:
    """Aggregate the video report per channel. Sorted by video count then views."""
    if not excel_file.exists():
        return []
    wb = load_workbook(excel_file, read_only=True)
    ws = wb.active
    agg = {}  # channel -> [count, views, likes, comments]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 9 or not row[1]:
            continue
        a = agg.setdefault(row[1], [0, 0, 0, 0])
        a[0] += 1
        a[1] += _num(row[6])
        a[2] += _num(row[7])
        a[3] += _num(row[8])
    wb.close()

    rows = []
    for ch, (cnt, views, likes, comments) in agg.items():
        eng = round((likes + comments) / views * 100, 2) if views else 0.0
        rows.append([ch, cnt, views, likes, comments, eng])
    rows.sort(key=lambda r: (-r[1], -r[2]))
    return rows


def write_summary(excel_file: Path, csv_path: Path, xlsx_path: Path) -> int:
    rows = build_rows(excel_file)
    if not rows:
        return 0
    totals = ["TỔNG", sum(r[1] for r in rows), sum(r[2] for r in rows),
              sum(r[3] for r in rows), sum(r[4] for r in rows), ""]

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        w.writerows(rows)
        w.writerow(totals)

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(HEADERS)
    for r in rows:
        ws.append(r)
    ws.append(totals)
    wb.save(xlsx_path)
    return len(rows)
