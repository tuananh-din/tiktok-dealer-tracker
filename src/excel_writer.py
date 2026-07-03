"""Append matched videos to an xlsx sheet; caller handles dedupe."""
import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook

HEADERS = [
    "Collected Date", "Channel", "Keyword", "Upload Date",
    "Video URL", "Caption", "Views", "Likes", "Comments",
]
_URL_COL = 4  # zero-based index of "Video URL" in a row


def _load_or_create(path: Path):
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "videos"
        ws.append(HEADERS)
    return wb, ws


def existing_urls(path: Path) -> set:
    """URLs already in the sheet, for cross-run dedupe."""
    if not path.exists():
        return set()
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) > _URL_COL and row[_URL_COL]:
            urls.add(row[_URL_COL])
    wb.close()
    return urls


def append_rows(path: Path, rows: list) -> int:
    """Append rows (list of dict). Returns number of rows written."""
    if not rows:
        return 0
    path.parent.mkdir(parents=True, exist_ok=True)
    wb, ws = _load_or_create(path)
    for r in rows:
        ws.append([
            r.get("collected_date"), r.get("channel"), r.get("keyword"),
            r.get("upload_date"), r.get("url"), r.get("caption"),
            r.get("views"), r.get("likes"), r.get("comments"),
        ])
    wb.save(path)
    return len(rows)


def export_csv(xlsx_path: Path, csv_path: Path) -> None:
    """Mirror the xlsx into a CSV (utf-8-sig so Excel + GitHub render it right)."""
    if not xlsx_path.exists():
        return
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
    wb.close()
