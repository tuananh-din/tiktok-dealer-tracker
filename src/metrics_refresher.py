#!/usr/bin/env python3
"""Refresh view/like/comment counts for already-saved videos.

The crawler captures a video ONCE (right after it is posted, when views are near
zero) and never re-fetches it — so metrics freeze at first-capture time while the
real video keeps gaining views. This module re-extracts metrics for recent videos
each run and updates the Excel report in place, then re-exports the CSV.

Only videos uploaded within a recent window are refreshed (older ones have
plateaued), keeping runs fast and polite to TikTok.
"""
import time
from pathlib import Path

from openpyxl import load_workbook

from src import excel_writer

# 1-based column positions in the videos sheet (see excel_writer.HEADERS).
_UP_COL, _URL_COL, _VIEW_COL, _LIKE_COL, _CMT_COL = 4, 5, 7, 8, 9


def _yyyymmdd(date_str: str) -> str:
    """'2026-07-04' -> '20260704'; passthrough if already compact/empty."""
    return (date_str or "").replace("-", "").strip()


def refresh_recent(excel_file, csv_file, since_yyyymmdd, extract_fn,
                   sleep_seconds=1.0, log=print) -> tuple:
    """Re-fetch metrics for saved videos uploaded on/after ``since_yyyymmdd``.

    extract_fn(url) -> metadata dict (as tiktok_crawler.extract_video) or None.
    Returns (checked, changed): how many recent videos were re-fetched and how
    many actually had different view counts.
    """
    excel_file = Path(excel_file)
    if not excel_file.exists():
        return (0, 0)

    wb = load_workbook(excel_file)
    ws = wb.active
    checked = changed = 0
    for row in ws.iter_rows(min_row=2):
        url = row[_URL_COL - 1].value
        if not url:
            continue
        up = _yyyymmdd(str(row[_UP_COL - 1].value or ""))
        if since_yyyymmdd and up and up < since_yyyymmdd:
            continue  # older than the window -> assume plateaued, skip

        checked += 1
        try:
            meta = extract_fn(url)
        except Exception as e:  # noqa: BLE001 - one bad video must not abort refresh
            log(f"    ! refresh failed {url}: {e}")
            continue
        if not meta or meta.get("views") is None:
            continue  # deleted/private/blocked -> keep the old numbers

        old_view = row[_VIEW_COL - 1].value
        row[_VIEW_COL - 1].value = meta["views"]
        row[_LIKE_COL - 1].value = meta["likes"]
        row[_CMT_COL - 1].value = meta["comments"]
        if str(old_view) != str(meta["views"]):
            changed += 1
        time.sleep(sleep_seconds)

    wb.save(excel_file)
    excel_writer.export_csv(excel_file, Path(csv_file))
    return (checked, changed)
